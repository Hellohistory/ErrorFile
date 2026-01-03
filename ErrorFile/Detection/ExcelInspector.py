# ErrorFile/Detection/ExcelInspector.py

import zipfile

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from xlrd.biffh import XLRDError

from ..report import (
    TAG_CORRUPTED,
    TAG_INVALID_FORMAT,
    TAG_IO_ERROR,
    fail_finding,
    ok_finding,
)


def _fast_check_xlsx(file_path):
    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            names = set(archive.namelist())
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required.issubset(names):
                return fail_finding(
                    "XLSX missing required parts; file may be corrupted.",
                    TAG_CORRUPTED,
                    TAG_INVALID_FORMAT,
                )
    except zipfile.BadZipFile as exc:
        return fail_finding(
            f"XLSX is not a valid ZIP container: {exc}",
            TAG_INVALID_FORMAT,
            error=str(exc),
        )
    except Exception as exc:
        return fail_finding(
            f"XLSX fast check failed: {exc}",
            TAG_IO_ERROR,
            error=str(exc),
        )
    return ok_finding("XLSX fast check passed.")


def check_excel_file(file_path, mode="deep"):
    """Check .xlsx integrity."""
    if mode == "fast":
        return _fast_check_xlsx(file_path)
    try:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                _ = sheet.cell(row=1, column=1).value
        finally:
            workbook.close()

        return ok_finding("XLSX deep check passed.")
    except InvalidFileException as exc:
        return fail_finding(
            f"XLSX corrupted or invalid: {exc}",
            TAG_CORRUPTED,
            TAG_INVALID_FORMAT,
            error=str(exc),
        )
    except Exception as exc:
        return fail_finding(
            f"XLSX deep check failed: {exc}",
            TAG_IO_ERROR,
            error=str(exc),
        )


def check_xls_file(file_path, mode="deep"):
    """Check .xls integrity."""
    try:
        xlrd.open_workbook(file_path)
        return ok_finding("XLS check passed.")
    except XLRDError as exc:
        return fail_finding(
            f"XLS corrupted or invalid: {exc}",
            TAG_CORRUPTED,
            TAG_INVALID_FORMAT,
            error=str(exc),
        )
    except Exception as exc:
        return fail_finding(
            f"XLS check failed: {exc}",
            TAG_IO_ERROR,
            error=str(exc),
        )
